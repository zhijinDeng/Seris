from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


SHEET_NAMES = {
    "overview": "说明与概览",
    "equipment_type": "设备类别",
    "equipment": "设备实例",
    "failure_mode": "失效模式",
    "work_order": "故障工单",
    "function": "功能关系",
    "relations": "关联概览",
    "cause": "原因类别",
    "dictionary": "数据字典",
}


def records(workbook, sheet_name: str) -> list[dict]:
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[4]]
    return [
        dict(zip(headers, [cell.value for cell in row], strict=False))
        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row)
    ]


def distribution(items: list[dict], field: str) -> dict[str, int]:
    counts = Counter(str(item.get(field)) for item in items)
    return dict(counts.most_common())


def main() -> None:
    parser = argparse.ArgumentParser(description="Profile the supplied masked SERES equipment dataset.")
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, data_only=True, read_only=False)
    missing_sheets = set(SHEET_NAMES.values()) - set(workbook.sheetnames)
    if missing_sheets:
        raise SystemExit(f"Missing sheets: {sorted(missing_sheets)}")

    equipment_types = records(workbook, SHEET_NAMES["equipment_type"])
    equipment = records(workbook, SHEET_NAMES["equipment"])
    failure_modes = records(workbook, SHEET_NAMES["failure_mode"])
    work_orders = records(workbook, SHEET_NAMES["work_order"])
    functions = records(workbook, SHEET_NAMES["function"])
    relation_rows = records(workbook, SHEET_NAMES["relations"])

    full_loop = [item for item in work_orders if item["structure_link_status"] == "结构链完整且类型一致"]
    linked_orders = [item for item in work_orders if item["has_valid_equipment_link"] is True]
    failure_linked = [item for item in work_orders if item["has_failure_mode_link"] is True]

    completeness_weight = {"高": 3, "中": 2, "低": 1}
    representative = sorted(
        full_loop,
        key=lambda item: (
            item["status_group"] == "已关闭",
            completeness_weight.get(item["completeness_grade"], 0),
            item["cause_category"] != "其他或未确认",
            item["action_category"] != "其他处置",
        ),
        reverse=True,
    )[:20]
    representative_fields = [
        "case_id",
        "effective_equipment_id",
        "effective_equipment_type_id",
        "equipment_family",
        "component_family",
        "source_group",
        "severity_band",
        "status_group",
        "phenomenon_category",
        "cause_category",
        "action_category",
        "outcome_group",
        "failure_mode_id",
        "completeness_grade",
        "privacy_group_size",
    ]

    output = {
        "source_file": args.workbook.name,
        "counts": {
            "equipment_types": len(equipment_types),
            "equipment_instances": len(equipment),
            "functions": len(functions),
            "failure_modes": len(failure_modes),
            "work_orders": len(work_orders),
        },
        "coverage": {
            "equipment_type_link": sum(item["has_type_link"] is True for item in equipment) / len(equipment),
            "equipment_knowledge_chain": sum(item["has_knowledge_chain"] is True for item in equipment) / len(equipment),
            "work_order_equipment_link": len(linked_orders) / len(work_orders),
            "work_order_failure_mode_link": len(failure_linked) / len(work_orders),
            "full_loop_candidate": len(full_loop) / len(work_orders),
        },
        "equipment_distributions": {
            field: distribution(equipment, field)
            for field in ["equipment_family", "asset_level_group", "status_group", "has_knowledge_chain", "work_order_count"]
        },
        "failure_mode_distributions": {
            field: distribution(failure_modes, field)
            for field in [
                "equipment_family",
                "component_family",
                "function_category",
                "phenomenon_category",
                "cause_category",
                "effect_category",
                "detection_category",
                "completeness_grade",
                "has_mechanism",
                "has_cause",
                "has_detection_method",
            ]
        },
        "work_order_distributions": {
            field: distribution(work_orders, field)
            for field in [
                "equipment_family",
                "component_family",
                "source_group",
                "severity_band",
                "status_group",
                "downtime_band",
                "phenomenon_category",
                "cause_category",
                "action_category",
                "outcome_group",
                "failure_mode_link_status",
                "structure_link_status",
                "relation_tier",
                "completeness_grade",
                "has_cause",
                "has_action",
            ]
        },
        "full_loop_distributions": {
            field: distribution(full_loop, field)
            for field in [
                "equipment_family",
                "source_group",
                "status_group",
                "phenomenon_category",
                "cause_category",
                "action_category",
                "outcome_group",
                "completeness_grade",
            ]
        },
        "representative_full_loop_candidates": [
            {field: item.get(field) for field in representative_fields} for item in representative
        ],
        "declared_relations": relation_rows,
    }
    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
